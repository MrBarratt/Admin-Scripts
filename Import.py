import os
import re
import csv
import concurrent.futures
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill

# Path to the folder containing the Excel files
folder_path = r'C:\Users\USER\OneDrive - XpertEase\Imports-Errors\Import\To be imported'
output_file_path = os.path.join(folder_path, "Duplicates remaining.txt")

# Create or clear the output file
with open(output_file_path, 'w') as output_file:
    output_file.write("Excel File Name, Remaining Rows\n")  # Header

# Function to extract the numbers from the filename
def extract_numbers_from_filename(filename):
    # Using regular expressions to find all numbers in the filename
    numbers = re.findall(r'\d+', filename)
    
    # Ensure there are at least 5 numbers in the filename
    if len(numbers) >= 5:
        first_number = numbers[0]  # 1st number (e.g., 20241021)
        second_number = numbers[1]  # 2nd number (e.g., 331)
        fourth_number = numbers[3]  # 4th number (e.g., 273)
        fifth_number = numbers[4]  # 5th number (e.g., 128)
        return second_number, fourth_number, fifth_number
    else:
        raise ValueError(f"Filename {filename} does not contain at least 5 numbers")

# Function to process a single file
def process_file(filename):
    file_path = os.path.join(folder_path, filename)
    print(f'Processing file: {file_path}')  # Print the full path

    try:
        # Load the workbook and iterate through each sheet
        workbook = load_workbook(file_path)

        for sheet in workbook.sheetnames:
            ws = workbook[sheet]

            # Define columns to clear formats
            columns_to_process = [chr(col) for col in range(ord('A'), ord('J') + 1)]  # Columns A to J

            for col in columns_to_process:
                # Clear formatting for the entire column
                for row in range(1, ws.max_row + 1):
                    cell = ws[f'{col}{row}']
                    cell.font = Font()  # Reset font
                    cell.fill = PatternFill()  # Reset fill
                    cell.border = Border()  # Clear borders
                    cell.alignment = Alignment()  # Reset alignment
                    cell.number_format = 'General'  # Reset number format
                    cell.protection = None  # Clear protection

            # Delete Row 1
            ws.delete_rows(1)

            # Extract numbers from the filename and fill Columns H, I, J
            try:
                second_number, fourth_number, fifth_number = extract_numbers_from_filename(filename)

                # Fill Columns H, I, and J with the extracted numbers
                for row in range(1, ws.max_row + 1):
                    ws[f'H{row}'].value = second_number  # Fill Column H with the 2nd number
                    ws[f'I{row}'].value = fourth_number  # Fill Column I with the 4th number
                    ws[f'J{row}'].value = fifth_number  # Fill Column J with the 5th number

            except ValueError as e:
                print(e)

            # Check if Column B contains any data
            column_b_has_data = any(ws[f'B{row}'].value for row in range(1, ws.max_row + 1))

            if column_b_has_data:
                # Format Column B to Custom Type = 0000000000
                for row in range(1, ws.max_row + 1):
                    cell = ws[f'B{row}']
                    cell.number_format = '0000000000'  # Set custom number format

                # Create Sheet 2 and copy Column B
                sheet2 = workbook.create_sheet(title="Sheet 2")
                for row in range(1, ws.max_row + 1):
                    cell_value = ws[f'B{row}'].value
                    sheet2[f'A{row}'] = cell_value  # Paste Column B into Column A of Sheet 2

                # Go back to Sheet 1 and format Column B as Text
                for row in range(1, ws.max_row + 1):
                    cell = ws[f'B{row}']
                    if cell.value is not None:
                        cell.value = str(cell.value).zfill(10)  # Convert to string and ensure 10 digits
                    cell.number_format = '@'  # Set the format of Column B to text

                # Copy values from Sheet 2's Column A to Sheet 1's Column B
                for row in range(1, ws.max_row + 1):
                    cell_value = sheet2[f'A{row}'].value
                    ws[f'B{row}'] = str(cell_value).zfill(10)  # Paste the value into Column B of Sheet 1 as text

                # Delete Sheet 2 after copying
                if "Sheet 2" in workbook.sheetnames:
                    std = workbook["Sheet 2"]
                    workbook.remove(std)  # Remove Sheet 2

            else:
                print(f'Column B is empty, skipping formatting for Column B in {filename}')

            # Delete duplicate rows based on Column A
            seen = set()
            rows_to_delete = []
            for row in range(1, ws.max_row + 1):
                cell_value = ws[f'A{row}'].value
                if cell_value in seen:
                    rows_to_delete.append(row)
                else:
                    seen.add(cell_value)

            # Delete rows in reverse order to avoid messing up the row indices
            for row in reversed(rows_to_delete):
                ws.delete_rows(row)

            # Count remaining rows after duplicates are removed
            remaining_rows = ws.max_row

        # Save the workbook after all operations
        workbook.save(file_path)
        print(f'Completed processing for {filename}')

        # Write the remaining row count to the output file
        with open(output_file_path, 'a') as output_file:
            output_file.write(f"{filename}, {remaining_rows}\n")

        # Save as CSV UTF-8 with "|" delimiter and " csv" appended to filename
        csv_file_name = filename.rsplit('.', 1)[0] + " csv.csv"  # Append " csv" and change extension
        csv_file_path = os.path.join(folder_path, csv_file_name)

        with open(csv_file_path, mode='w', newline='', encoding='utf-8') as csv_file:
            writer = csv.writer(csv_file, delimiter='|')  # Use "|" as the delimiter
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        print(f'Saved {csv_file_name} as a UTF-8 CSV file with "|" delimiter.')

    except PermissionError:
        print(f"Permission denied for file: {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

# Run the processing without a timeout
with concurrent.futures.ThreadPoolExecutor() as executor:
    futures = {executor.submit(process_file, filename): filename for filename in os.listdir(folder_path)
               if filename.endswith('.xlsx') or filename.endswith('.xlsm')}

    for future in concurrent.futures.as_completed(futures):
        filename = futures[future]
        try:
            future.result()
        except Exception as e:
            print(f"An error occurred while processing {filename}: {e}")

print('All operations completed for all Excel files in the specified folder.')
